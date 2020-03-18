from openpyxl import Workbook
import openpyxl
arquivo_excel = Workbook()

planilha1 = arquivo_excel.active

#Alterar nome da Planilha
planilha1.title = 'Gastos'

# Criar Novas Planilhas no mesmo Arquivo
planilha2 = arquivo_excel.create_sheet("Ganhos")
planilha3 = arquivo_excel.create_sheet("PPP")

'''O atributo sheetnames do workbook guarda em uma lista os nomes das planilhas criadas exatamente na mesma ordem q sao criadas.
'''

#verificar em q ordem as planilhas estao criadas
print( arquivo_excel.sheetnames)

# Adicionar Dados a Planilha

# Forma 1:

planilha1['A1'] = 'Categorias'
planilha1['B1'] = 'Valor'
planilha1['A1'] = 'Restaurante'
planilha1['B2'] = '45.99'


# Forma 2:
#escrever linha a linha usando tuplas

valores = [
    ("Categoria", "Valor"),
    ("Restaurante", 45.99),
    ("Transporte", 208.45),
    ("Viagem", 558.54)
    ]
for linha in valores:
    planilha2.append(linha)


# Forma 3

# Utilizando o metodo cell
planilha3.cell(row=3, column=1,
              value= 35.99)

# Adicionar formulas a Planilha

planilha1['C1'] = '=SUM(23,5)'

# Ler dados da Planilha

c1 = planilha1['C1']
#Imprime o valor da celula C1
print(c1.value)
a1 = planilha1.cell(column=1, row = 1)
#Imprimir o valor da celula a1
print(a1.value)

#Forma 2 de ler ficheiros

max_linha = planilha1.max_row
max_coluna = planilha1.max_column
for i in range(1, max_linha + 1):
    for j in range(1, max_coluna + 1):
        print(planilha1.cell(row=i,
   column= j).value, end ="-")

# Salvando Arquivo
arquivo_excel.save("Relatorio.xlsx")

#Carregando uma Planilha existente

from openpyxl import load_workbook
#caminho = 'c:/users/ate/o/seu/arquivo.xlsx'
caminho ='C:Relatorio.xlsx'
#C:\Users\Leonel K4T\OneDrive\deV\Py\Excel Thang\pyExcel
arquivo_excel = load_workbook(caminho)


#Copiando dados entre planilhas

'''
original = arquivo_excel.get_sheet_by_name('Gastos')
copia = arquivo_excel.copy_worksheet(copia)
arquivo_excel.save('Relatorio.xlsx')'''














